import io


from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file,  make_response
from openpyxl import Workbook
from models.type_document import Type_document
from models.user import User
from utils.db import db


users = Blueprint("users", __name__)


@users.route('/')
def index():
    types_document = Type_document.query.all()
    return render_template('index.html', types_document=types_document)


@users.route('/get_user/>', methods=['GET', 'POST'])
def get_user_by_ID():
    number_ID = request.form['number_ID']

    type_document =int(request.form['type_document'])


    if request.method == 'POST':
        types_document = Type_document.query.all()


        users = User.query.filter_by(number_ID=number_ID).all()
        if not users:
            return 'El número de identificación no existe en la base de datos.'
        
        if users[0].type_id != type_document:
            return 'El número de identificación pertenece a un usuario con un tipo de documento diferente.'


        user = User.query.filter_by(number_ID=number_ID).first()
        
        if user:
            user_info = {
                'number_ID': user.number_ID,
                'name': user.name,
                'last_name':user.last_name,
                'email': user.email,
                'phone': user.phone,
                'type_id':user.type_id
            }
            return render_template('index.html', user_info = user_info, types_document=types_document)
        else:
            return ('No se encontraron resultados.')

    return render_template('index.html')


@users.route('/export_user')
def export_user():
    number_ID = request.args.get('number_ID')
    print(number_ID)
    user  = User.query.filter_by(number_ID=number_ID).first()
    print(user)
    
    if user is None:
        return 'No se encontró ningún usuario con ese número de identificación'
    
    workbook = Workbook()
    

    sheet = workbook.create_sheet('Usuario')
    
    
    sheet.cell(row=1, column=1, value='Nombre')
    sheet.cell(row=1, column=2, value='Apellido')
    sheet.cell(row=1, column=3, value='Correo electrónico')
    sheet.cell(row=1, column=4, value='Teléfono')
    sheet.cell(row=1, column=5, value='Tipo de documento')

    sheet.cell(row=2, column=1, value=user.name)
    sheet.cell(row=2, column=2, value=user.last_name)
    sheet.cell(row=2, column=3, value=user.email)
    sheet.cell(row=2, column=4, value=user.phone)
    sheet.cell(row=2, column=5, value=user.type_id)
    
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    

    response = make_response(buffer.getvalue())
    filename = f'{user.name} {user.last_name}.xlsx'
    response.headers.set('Content-Disposition', f'attachment; filename="{filename}"')
    

    return response